from openpyxl import Workbook 
from openpyxl.styles import Font 
import pandas as pd
wb = Workbook() 
sheet = wb.active
sheet.title = "studentdata"
Name = ["RAM", "JANU", "ARYA"]
College = ["JIT", "RNSIT", "SJBIT"]
Marks = ["89", "95", "98"]
sheet.cell(row = 1, column = 1).value = "Name" 
sheet.cell(row = 1, column = 2).value = "College" 
sheet.cell(row = 1, column = 3).value = "Marks" 
ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft
for i in range(2,5):
    sheet.cell(row = i, column = 1).value = Name[i-2]
    sheet.cell(row = i, column = 2).value = College[i-2]
    sheet.cell(row = i, column = 3).value = Marks[i-2]
wb.save("9bprogram.xlsx")